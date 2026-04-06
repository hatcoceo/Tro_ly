import os
import subprocess
import datetime
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

from openpyxl import load_workbook


class PlotHandler:
    def __init__(self, assistant):
        self.assistant = assistant

    def can_handle(self, command: str) -> bool:
        return command.startswith("plot ")

    def handle(self, command: str) -> None:
        try:
            print("📥 Lệnh:", command)

            content = command.replace("plot ", "").strip()

            # =========================
            # 📊 EXCEL
            # =========================
            if content.startswith("excel "):
                parts = content.split()

                if len(parts) < 5:
                    raise ValueError("Cú pháp: plot excel file.xlsx <cột|header> start_row end_row")

                _, file_path, col_or_header, start_row, end_row = parts

                start_row = int(start_row)
                end_row = int(end_row)

                numbers = self.read_excel_column(
                    file_path,
                    col_or_header,
                    start_row,
                    end_row
                )

                folder = self.build_folder("excel", file_path)
                self.plot(numbers, title=f"{file_path}_{col_or_header}", folder=folder)
                return

            # =========================
            # 📂 FILE TEXT
            # =========================
            if os.path.exists(content):
                print("📂 Đọc file:", content)

                with open(content, "r", encoding="utf-8") as f:
                    data = f.read()

                numbers = self.parse_numbers(data)

                folder = self.build_folder("txt", content)
                self.plot(numbers, title=content, folder=folder)
                return

            # =========================
            # ✏️ INPUT TRỰC TIẾP
            # =========================
            print("✏️ Input:", content)
            numbers = self.parse_numbers(content)

            print("🔢 Parsed:", numbers)

            folder = self.build_folder("manual", "input")
            self.plot(numbers, title="manual_input", folder=folder)

        except Exception as e:
            print(f"⚠️ Lỗi: {e}")

    # =========================
    # 📂 TẠO THƯ MỤC
    # =========================
    def build_folder(self, category, name):
        base = "plots"

        safe_name = os.path.splitext(os.path.basename(name))[0]
        safe_name = safe_name.replace(" ", "_")

        folder = os.path.join(base, category, safe_name)
        os.makedirs(folder, exist_ok=True)

        return folder

    # =========================
    # 📊 ĐỌC EXCEL
    # =========================
    def read_excel_column(self, file_path, col_or_header, start_row, end_row):
        if not os.path.exists(file_path):
            raise FileNotFoundError("Không tìm thấy file Excel")

        wb = load_workbook(file_path)
        ws = wb.active

        col = None
        header_row = 1

        # 🔍 tìm header
        for cell in ws[header_row]:
            if cell.value:
                name = str(cell.value).strip().lower()
                if name == col_or_header.strip().lower():
                    col = cell.column_letter
                    print(f"🔎 Header '{col_or_header}' → cột {col}")
                    break

        # fallback: cột A/B/C
        if col is None:
            if col_or_header.isalpha():
                col = col_or_header.upper()
                print(f"📌 Dùng cột: {col}")
            else:
                raise ValueError("Không tìm thấy cột hợp lệ")

        numbers = []

        for row in range(start_row, end_row + 1):
            value = ws[f"{col}{row}"].value

            if value is None:
                continue

            try:
                numbers.append(float(value))
            except:
                pass

        if not numbers:
            raise ValueError("Không có dữ liệu số")

        print("📊 Excel:", numbers)

        return numbers

    # =========================
    # 🔢 PARSE SỐ
    # =========================
    def parse_numbers(self, text: str):
        text = text.replace(",", " ").replace("\n", " ")
        parts = text.split()

        numbers = []
        for p in parts:
            try:
                numbers.append(float(p))
            except:
                pass

        if not numbers:
            raise ValueError("Không tìm thấy số")

        return numbers

    # =========================
    # 📈 VẼ BIỂU ĐỒ + LƯU
    # =========================
    def plot(self, numbers, title="plot", folder="plots"):
        x = list(range(1, len(numbers) + 1))

        plt.figure(figsize=(10, 6))
        plt.plot(x, numbers, marker='o', label="Dữ liệu")

        # =========================
        # 📊 TRUNG BÌNH (bỏ số cuối)
        # =========================
        avg_data = numbers[:-1] if len(numbers) > 1 else numbers
        avg = sum(avg_data) / len(avg_data)

        # =========================
        # 🔴 ĐƯỜNG TRUNG BÌNH
        # =========================
        plt.axhline(
            y=avg,
            color='red',
            linestyle='--',
            linewidth=2,
            label=f"Trung bình (bỏ số cuối) = {avg:.2f}"
        )
        # 🔵 ĐƯỜNG N1
        n1 = numbers[0]
        plt.axhline(
            y=n1,
            color='blue',
            linestyle='-',
            linewidth=2,
            label=f"N1 = {n1:.2f}"
        )

        # =========================
        # 🔴 ĐIỂM TRUNG BÌNH
        # =========================
        mid_x = (len(numbers) + 1) / 2
        plt.scatter(
            mid_x,
            avg,
            color='red',
            s=120,
            zorder=5,
            edgecolors='black'
        )

        # =========================
        # 🏷️ TEXT HIỂN THỊ
        # =========================
        plt.text(
            mid_x,
            avg,
            f"{avg:.2f}",
            ha='center',
            va='bottom',
            fontsize=10,
            bbox=dict(facecolor='white', alpha=0.7)
        )

        # =========================
        # 🎨 HIỂN THỊ
        # =========================
        plt.title(title)
        plt.xlabel("N")
        plt.ylabel("P")
        plt.grid(True)
        plt.legend(loc='best')

        # =========================
        # 💾 LƯU FILE
        # =========================
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

        safe_title = title.replace(" ", "_").replace("|", "").replace("/", "_")

        filename = f"{safe_title}_{timestamp}.png"
        file_path = os.path.join(folder, filename)

        plt.tight_layout()
        plt.savefig(file_path, bbox_inches='tight', dpi=150)
        plt.close()

        print(f"📊 Đã lưu: {file_path}")

        self.open_image(file_path)

    # =========================
    # 🖼️ MỞ ẢNH
    # =========================
    def open_image(self, path: str):
        try:
            subprocess.run(["xdg-open", path])
        except Exception:
            print("⚠️ Không mở được ảnh")
            print("👉 Mở thủ công:", path)


def register(assistant):
    assistant.handlers.append(PlotHandler(assistant))


plugin_info = {
    "enabled": True,
    "register": register,
    "command_handle": [
        "plot 1 2 3",
        "plot data.txt",
        "plot excel file.xlsx A 2 10",
        "plot excel file.xlsx Price 2 10"
    ]
}