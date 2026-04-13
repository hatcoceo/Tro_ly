import os
import csv
from typing import Any, List
try:
    from openpyxl import Workbook
except ImportError:
    raise ImportError('Vui lòng cài đặt openpyxl: pip install openpyxl')
    
plugin_info = {
    'enabled': True,
    'register': None,
    'command_handle': [
        'txt2excel',
        'convert txt to excel'
    ]
}
class TxtToExcelHandler:
    """Xử lý lệnh chuyển đổi file txt sang excel"""

    def can_handle(self, command: str) ->bool:
        """Kiểm tra lệnh có phải là chuyển đổi txt->excel không"""
        cmd = command.lower().strip()
        return cmd in ('txt2excel', 'convert txt to excel')

    def handle(self, command: str) ->None:
        """Thực hiện chuyển đổi file txt sang excel"""
        print(
            '\n📄 Chuyển đổi file TXT (cấu trúc: time open high low close volume) sang Excel'
            )
        txt_path = input('Nhập đường dẫn file TXT: ').strip()
        if not txt_path:
            print('❌ Bạn chưa nhập đường dẫn.')
            return
        if not os.path.exists(txt_path):
            print(f'❌ Không tìm thấy file: {txt_path}')
            return
        default_excel = os.path.splitext(txt_path)[0] + '.xlsx'
        excel_path = input(
            f'Nhập đường dẫn file Excel (Enter để dùng {default_excel}): '
            ).strip()
        if not excel_path:
            excel_path = default_excel
        if not excel_path.lower().endswith('.xlsx'):
            excel_path += '.xlsx'
        data_rows = []
        try:
            with open(txt_path, 'r', encoding='utf-8') as f:
                header_line = f.readline()
                if not header_line:
                    print('❌ File TXT trống.')
                    return
                headers = header_line.strip().split()
                expected_headers = ['time', 'open', 'high', 'low', 'close',
                    'volume']
                if len(headers) != 6 or any(h not in expected_headers for h in
                    headers):
                    print(
                        f'⚠️ Cảnh báo: Header không đúng chuẩn. Tìm thấy: {headers}'
                        )
                    print(
                        '   Dữ liệu vẫn được chuyển nhưng cột có thể không chính xác.'
                        )
                reader = csv.reader(f, delimiter=' ', skipinitialspace=True)
                for row_num, row in enumerate(reader, start=2):
                    if not row or not any(field.strip() for field in row):
                        continue
                    if len(row) < 6:
                        row.extend([''] * (6 - len(row)))
                    data_rows.append(row)
        except Exception as e:
            print(f'❌ Lỗi khi đọc file TXT: {e}')
            return
        if not data_rows:
            print('❌ Không có dữ liệu hợp lệ trong file TXT.')
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Data'
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)
            for row_idx, row_data in enumerate(data_rows, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    if col_idx > 1:
                        try:
                            val_clean = value.replace(',', '')
                            numeric_val = float(val_clean)
                            ws.cell(row=row_idx, column=col_idx, value=
                                numeric_val)
                        except ValueError:
                            ws.cell(row=row_idx, column=col_idx, value=value)
                    else:
                        ws.cell(row=row_idx, column=col_idx, value=value)
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[col_letter].width = adjusted_width
            wb.save(excel_path)
            print(
                f'✅ Chuyển đổi thành công! File Excel đã được lưu tại: {excel_path}'
                )
            print(f'   Số dòng dữ liệu: {len(data_rows)}')
        except Exception as e:
            print(f'❌ Lỗi khi ghi file Excel: {e}')


def register(assistant: Any) ->None:
    """Đăng ký handler vào trợ lý"""
    handler = TxtToExcelHandler()
    assistant.handlers.append(handler)
    print(
        "🔌 Plugin txt2excel đã được nạp. Gõ 'txt2excel' để chuyển đổi file TXT sang Excel."
        )


plugin_info['register'] = register
