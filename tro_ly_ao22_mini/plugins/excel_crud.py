# thiết kế lại kiến để dễ dàng mở rộng 
# https://chat.deepseek.com/share/1ug4jod6noos7fk8om
# thêm chức năng 
# thêm chức năng ghi công thức cho ô
import os
import shlex
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
import matplotlib.pyplot as plt

def with_worksheet(func):
    """
    Decorator tự động:
    - Load workbook từ self.file (nếu chưa có thì báo lỗi)
    - Lấy active worksheet
    - Gọi func với tham số ws đầu tiên
    - Tự động save workbook sau khi func chạy (nếu không có lỗi)
    """
    def wrapper(self, *args, **kwargs):
        if not self.file:
            raise Exception("Chưa chỉ định file. Dùng -f <file> hoặc lệnh setfile")
        if not os.path.exists(self.file):
            raise Exception(f"File không tồn tại: {self.file}")
        wb = load_workbook(self.file)
        ws = wb.active
        try:
            result = func(self, ws, *args, **kwargs)
            wb.save(self.file)
            return result
        except Exception as e:
            # Không save nếu có lỗi
            raise e
    return wrapper


class ExcelProHandler:
    """
    Xử lý lệnh excel. Mọi chức năng đều là method có tên cmd_<tên_lệnh>.
    Tự động đăng ký khi khởi tạo.
    """

    def __init__(self, assistant):
        self.assistant = assistant
        self.file = None          # file mặc định
        self.commands = {}        # registry: tên lệnh -> method

        # Tự động đăng ký tất cả method bắt đầu bằng 'cmd_'
        for attr_name in dir(self):
            if attr_name.startswith('cmd_'):
                cmd_name = attr_name[4:]    # bỏ 'cmd_'
                method = getattr(self, attr_name)
                self.commands[cmd_name] = method

        # Thêm alias (có thể mở rộng)
        self.aliases = {
            'delrows': 'delete_rows_range',
            'delcolrange': 'delete_column_range',
            'avg_range': 'avg_range',
        }
        # Đăng ký alias
        for alias, target in self.aliases.items():
            if target in self.commands:
                self.commands[alias] = self.commands[target]

    def can_handle(self, command: str) -> bool:
        return command.startswith('excel')

    def handle(self, command: str):
        try:
            # Dùng shlex để parse chuỗi có dấu ngoặc kép
            parts = shlex.split(command)
            if len(parts) < 2:
                print("❌ Lệnh excel thiếu tham số")
                return

            # Bỏ qua 'excel' ở đầu
            args = parts[1:]

            # Xử lý flag -f / --file
            new_args = []
            filename = None
            i = 0
            while i < len(args):
                if args[i] in ('-f', '--file'):
                    if i + 1 < len(args):
                        filename = args[i + 1]
                        i += 2
                        continue
                    else:
                        print("⚠️ Thiếu tên file sau -f/--file")
                        return
                new_args.append(args[i])
                i += 1

            if filename:
                self.file = filename
            elif self.file is None:
                print("⚠️ Chưa chỉ định file. Dùng -f <tên_file> hoặc lệnh setfile")
                return

            if not new_args:
                print("❌ Thiếu tên lệnh")
                return

            cmd = new_args[0].lower()
            cmd_args = new_args[1:] if len(new_args) > 1 else []

            # Tìm method trong registry
            method = self.commands.get(cmd)
            if not method:
                print(f"❌ Lệnh không hợp lệ: {cmd}")
                return

            # Gọi method, có thể có hoặc không decorator @with_worksheet
            # Một số lệnh không cần worksheet (create, setfile, copy, manual...)
            method(cmd_args)

        except Exception as e:
            print(f"⚠️ Lỗi: {e}")

    # ================== ĐỊNH NGHĨA CÁC LỆNH ==================
    # Mỗi lệnh là method cmd_<tên>, với tham số args (list)
    # Nếu lệnh cần worksheet, hãy dùng decorator @with_worksheet
    # và tham số đầu tiên là ws (worksheet)

    def cmd_create(self, args):
        """Tạo file Excel mới"""
        wb = Workbook()
        wb.save(self.file)
        print(f"✅ Đã tạo file {self.file}")

    @with_worksheet
    def cmd_add(self, ws, args):
        """Thêm dòng dữ liệu"""
        if not args:
            print("⚠️ excel add <giá_trị1> <giá_trị2> ...")
            return
        row = []
        for x in args:
            try:
                row.append(float(x))
            except:
                row.append(x)
        ws.append(row)
        print(f"✅ Đã thêm: {row}")

    @with_worksheet
    def cmd_read(self, ws, args):
        """Đọc nội dung"""
        for row in ws.iter_rows(values_only=True):
            print(row)

    @with_worksheet
    def cmd_update(self, ws, args):
        """Cập nhật ô: update <hàng> <cột> <giá_trị>"""
        if len(args) < 3:
            print("⚠️ excel update <hàng> <cột> <giá_trị>")
            return
        r, c = int(args[0]), int(args[1])
        val = args[2]
        ws.cell(row=r, column=c).value = val
        print("✅ Đã cập nhật")

    @with_worksheet
    def cmd_delete(self, ws, args):
        """Xóa dòng: delete <hàng>"""
        if not args:
            print("⚠️ excel delete <hàng>")
            return
        ws.delete_rows(int(args[0]))
        print("🗑 Đã xóa dòng")

    @with_worksheet
    def cmd_delrows(self, ws, args):
        """Xóa khoảng dòng: delrows <hàng_đầu> <hàng_cuối>"""
        if len(args) < 2:
            print("⚠️ excel delrows <hàng_đầu> <hàng_cuối>")
            return
        start, end = int(args[0]), int(args[1])
        count = end - start + 1
        ws.delete_rows(start, count)
        print(f"🗑 Đã xóa dòng {start} đến {end}")

    @with_worksheet
    def cmd_delcolrange(self, ws, args):
        """Xóa dữ liệu cột theo dòng: delcolrange <cột> <hàng_đầu> <hàng_cuối>"""
        if len(args) < 3:
            print("⚠️ excel delcolrange <cột> <hàng_đầu> <hàng_cuối>")
            return
        col, start, end = int(args[0]), int(args[1]), int(args[2])
        for row in range(start, end + 1):
            ws.cell(row=row, column=col).value = None
        print(f"🗑 Đã xóa dữ liệu cột {col}, dòng {start}-{end}")

    @with_worksheet
    def cmd_find(self, ws, args):
        """Tìm kiếm từ khóa"""
        if not args:
            print("⚠️ excel find <từ_khóa>")
            return
        keyword = args[0]
        found = False
        for row in ws.iter_rows(values_only=True):
            if any(keyword in str(cell) for cell in row):
                print("🔍", row)
                found = True
        if not found:
            print(f"Không tìm thấy '{keyword}'")

    @with_worksheet
    def cmd_avg(self, ws, args):
        """Tính trung bình tất cả số"""
        nums = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, (int, float)):
                    nums.append(cell)
        if nums:
            avg = sum(nums) / len(nums)
            self.assistant.context['avg'] = avg
            print(f"📊 AVG = {avg}")
        else:
            print("⚠️ Không có số nào")

    @with_worksheet
    def cmd_avg_range(self, ws, args):
        """Trung bình cột theo dòng: avg_range <cột> <hàng_đầu> <hàng_cuối>"""
        if len(args) < 3:
            print("⚠️ excel avg_range <cột> <hàng_đầu> <hàng_cuối>")
            return
        col, start, end = int(args[0]), int(args[1]), int(args[2])
        nums = []
        for row in range(start, end + 1):
            val = ws.cell(row=row, column=col).value
            try:
                nums.append(float(val))
            except:
                continue
        if nums:
            avg = sum(nums) / len(nums)
            print(f"📊 AVG (cột {col}, dòng {start}-{end}) = {avg}")
            self.assistant.context['avg_range'] = avg
        else:
            print("⚠️ Không có dữ liệu số trong khoảng")

    @with_worksheet
    def cmd_chart(self, ws, args):
        """Vẽ biểu đồ"""
        data = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, (int, float)):
                    data.append(cell)
        if not data:
            print("⚠️ Không có dữ liệu số")
            return
        plt.figure()
        plt.plot(data)
        plt.title("Excel Data Chart")
        plt.savefig("chart.png")
        plt.close()
        print("📈 Đã lưu chart.png")

    @with_worksheet
    def cmd_auto(self, ws, args):
        """Quyết định BUY/WAIT dựa trên số cuối"""
        data = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, (int, float)):
                    data.append(cell)
        if len(data) < 2:
            print("⚠️ Không đủ dữ liệu (cần ít nhất 2 số)")
            return
        avg_prev = sum(data[:-1]) / len(data[:-1])
        last = data[-1]
        print(f"Trung bình (trừ số cuối): {avg_prev}")
        print(f"Số cuối: {last}")
        decision = "BUY" if last < avg_prev else "WAIT"
        print(f"🤖 Quyết định: {decision}")
        self.assistant.context['decision'] = decision

    def cmd_copy(self, args):
        """Sao chép file: copy [nguồn] đích"""
        if len(args) == 1:
            source = self.file
            dest = args[0]
        elif len(args) == 2:
            source, dest = args[0], args[1]
        else:
            print("⚠️ excel copy <đích>  hoặc  excel copy <nguồn> <đích>")
            return
        if not os.path.exists(source):
            print(f"⚠️ File nguồn không tồn tại: {source}")
            return
        wb = load_workbook(source)
        wb.save(dest)
        print(f"✅ Đã sao chép {source} -> {dest}")

    def cmd_setfile(self, args):
        """Đặt file mặc định: setfile <tên_file>"""
        if not args:
            print("⚠️ excel setfile <tên_file>")
            return
        self.file = args[0]
        print(f"📁 Đã đặt file mặc định: {self.file}")

    @with_worksheet
    def cmd_header(self, ws, args):
        """
        Sửa hàng tiêu đề (dòng 1)
        Cú pháp:
            header cột "nội dung"
            header "nd1" "nd2" ...
        """
        if not args:
            print("⚠️ excel header <cột> \"nội dung\"  hoặc  excel header \"nd1\" \"nd2\" ...")
            return
        if len(args) == 2 and args[0].isdigit():
            col = int(args[0])
            value = args[1].strip('"')
            ws.cell(row=1, column=col).value = value
            print(f"✅ Đã sửa tiêu đề cột {col} thành: {value}")
        else:
            for idx, val in enumerate(args, start=1):
                clean_val = val.strip('"')
                ws.cell(row=1, column=idx).value = clean_val
            print(f"✅ Đã cập nhật {len(args)} ô tiêu đề")

    @with_worksheet
    def cmd_comment(self, ws, args):
        """Thêm comment: comment <hàng> <cột> \"nội dung\" """
        if len(args) < 3:
            print("⚠️ excel comment <hàng> <cột> \"nội dung\"")
            return
        row, col = int(args[0]), int(args[1])
        comment_text = ' '.join(args[2:]).strip('"')
        cell = ws.cell(row=row, column=col)
        if cell.comment:
            cell.comment = None
        cell.comment = Comment(comment_text, "User")
        print(f"✅ Đã thêm comment vào ô ({row},{col}): \"{comment_text}\"")

    def cmd_manual(self, args):
        """Nhập dữ liệu thủ công tương tác"""
        if not self.file:
            print("⚠️ Chưa chỉ định file. Dùng setfile hoặc -f trước.")
            return
        # Tạo file nếu chưa tồn tại
        if not os.path.exists(self.file):
            print(f"⚠️ File {self.file} chưa tồn tại, sẽ tạo mới.")
            wb = Workbook()
            wb.save(self.file)
        wb = load_workbook(self.file)
        ws = wb.active
        try:
            num_cols = int(input("Nhập số cột dữ liệu: "))
        except:
            print("❌ Số cột không hợp lệ")
            return
        print("Nhập dữ liệu từng dòng (cách nhau bằng khoảng trắng hoặc dấu phẩy), 'done' để kết thúc")
        row_count = 0
        while True:
            line = input(f"Dòng {row_count+1}: ").strip()
            if line.lower() == 'done':
                break
            if line.lower() == 'show':
                for r in ws.iter_rows(values_only=True):
                    print(r)
                continue
            parts = line.replace(',', ' ').split()
            if len(parts) != num_cols:
                print(f"⚠️ Cần {num_cols} giá trị, bạn nhập {len(parts)}. Thử lại.")
                continue
            row_data = []
            for p in parts:
                try:
                    row_data.append(float(p))
                except:
                    row_data.append(p)
            ws.append(row_data)
            row_count += 1
            print(f"✅ Đã thêm dòng {row_count}")
        if row_count > 0:
            wb.save(self.file)
            print(f"💾 Đã lưu {row_count} dòng vào {self.file}")
        else:
            print("Không có dữ liệu nào được thêm.")

    @with_worksheet
    def cmd_delcol(self, ws, args):
        """Xóa hẳn một cột: delcol <cột>"""
        if not args:
            print("⚠️ excel delcol <cột>")
            return
        col = int(args[0])
        ws.delete_cols(col)
        print(f"🗑 Đã xóa cột {col}")

    @with_worksheet
    def cmd_autofit(self, ws, args):
        """
        Tự động khớp độ rộng cột
        autofit              -> tất cả cột có dữ liệu
        autofit 1 3 5        -> chỉ các cột 1,3,5
        """
        if args:
            try:
                columns = [int(a) for a in args]
            except:
                print("⚠️ Tham số cột phải là số")
                return
        else:
            # Xác định cột cuối cùng có dữ liệu
            max_col = 0
            for row in ws.iter_rows(values_only=True):
                for idx, cell in enumerate(row, start=1):
                    if cell is not None and idx > max_col:
                        max_col = idx
            if max_col == 0:
                print("⚠️ Không có dữ liệu")
                return
            columns = list(range(1, max_col + 1))

        for col in columns:
            max_length = 0
            col_letter = ws.cell(row=1, column=col).column_letter
            for row in range(1, ws.max_row + 1):
                val = ws.cell(row=row, column=col).value
                if val is not None:
                    length = len(str(val))
                    if length > max_length:
                        max_length = length
            adjusted_width = min(max(max_length + 2, 8), 50)
            ws.column_dimensions[col_letter].width = adjusted_width
        print(f"✅ Đã tự động khớp độ rộng cho {len(columns)} cột")

    @with_worksheet
    def cmd_colorminmax(self, ws, args):
        """
        Tô màu min (xanh) và max (hồng) trong một cột (bỏ qua dòng 1)
        colorminmax <cột>
        """
        if not args:
            print("⚠️ excel colorminmax <cột>")
            return
        col = int(args[0])
        min_val = None
        max_val = None
        min_cells = []
        max_cells = []
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            val = cell.value
            if isinstance(val, (int, float)):
                if min_val is None or val < min_val:
                    min_val = val
                    min_cells = [cell]
                elif val == min_val:
                    min_cells.append(cell)
                if max_val is None or val > max_val:
                    max_val = val
                    max_cells = [cell]
                elif val == max_val:
                    max_cells.append(cell)
        if min_val is None:
            print("⚠️ Không có dữ liệu số trong cột")
            return
        min_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        max_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
        for cell in min_cells:
            cell.fill = min_fill
        for cell in max_cells:
            cell.fill = max_fill
        print(f"✅ Đã tô màu min={min_val} (xanh) và max={max_val} (hồng) tại cột {col}")

    # ================== HƯỚNG DẪN THÊM LỆNH MỚI ==================
    # Để thêm chức năng mới, chỉ cần viết method cmd_<tên_lệnh>
    # - Nếu cần thao tác với worksheet, dùng decorator @with_worksheet
    # - Tham số args là list các string (đã được split và xử lý quote)
    # - Tự động đăng ký vào self.commands nhờ __init__
    # - Có thể thêm alias bằng cách thêm vào dict self.aliases
    # Ví dụ:
    # @with_worksheet
    # def cmd_hello(self, ws, args):
    #     print("Hello world!", args)
    # Người dùng gõ: excel hello 1 2 3

    # ================== THÊM CÁC CHỨC NĂNG QUAN TRỌNG KHÁC ==================
    
    @with_worksheet
    def cmd_find_replace(self, ws, args):
        """
        Tìm và thay thế chuỗi trong toàn bộ worksheet.
        Cú pháp: excel find_replace "tìm" "thay"
        """
        if len(args) < 2:
            print("⚠️ excel find_replace \"từ_cần_tìm\" \"thay_thế\"")
            return
        find_str = args[0].strip('"')
        replace_str = args[1].strip('"')
        count = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and find_str in cell.value:
                    cell.value = cell.value.replace(find_str, replace_str)
                    count += 1
        print(f"✅ Đã thay thế '{find_str}' → '{replace_str}' trong {count} ô")
    
    
    @with_worksheet
    def cmd_remove_empty_rows(self, ws, args):
        """
        Xóa tất cả các dòng hoàn toàn trống.
        Cú pháp: excel remove_empty_rows
        """
        rows_to_delete = []
        for row_idx in range(1, ws.max_row + 1):
            is_empty = True
            for col_idx in range(1, ws.max_column + 1):
                if ws.cell(row=row_idx, column=col_idx).value is not None:
                    is_empty = False
                    break
            if is_empty:
                rows_to_delete.append(row_idx)
        # Xóa từ dưới lên để tránh ảnh hưởng chỉ số
        for row_idx in reversed(rows_to_delete):
            ws.delete_rows(row_idx)
        print(f"🗑 Đã xóa {len(rows_to_delete)} dòng trống")
    
    
    @with_worksheet
    def cmd_sort(self, ws, args):
        """
        Sắp xếp dữ liệu theo một cột (có header dòng 1).
        Cú pháp: excel sort <cột> [asc|desc]
        Mặc định: asc (tăng dần)
        """
        if not args:
            print("⚠️ excel sort <cột> [asc|desc]")
            return
        try:
            col = int(args[0])
        except ValueError:
            print("⚠️ Cột phải là số")
            return
        order = args[1].lower() if len(args) > 1 else 'asc'
        reverse = (order == 'desc')
        
        # Lấy dữ liệu từ dòng 2 trở đi (bỏ qua header)
        data_rows = []
        for row in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=row, column=col).value
            row_data = [ws.cell(row=row, column=c).value for c in range(1, ws.max_column + 1)]
            data_rows.append((cell_val, row_data))
        
        # Sắp xếp
        try:
            data_rows.sort(key=lambda x: x[0] if x[0] is not None else '', reverse=reverse)
        except TypeError:
            # Nếu có mixed types, ép về string để so sánh
            data_rows.sort(key=lambda x: str(x[0]) if x[0] is not None else '', reverse=reverse)
        
        # Ghi lại dữ liệu
        for new_row_idx, (_, row_data) in enumerate(data_rows, start=2):
            for col_idx, val in enumerate(row_data, start=1):
                ws.cell(row=new_row_idx, column=col_idx).value = val
        print(f"✅ Đã sắp xếp theo cột {col} ({order})")
    
    
    @with_worksheet
    def cmd_stat(self, ws, args):
        """
        Thống kê cột (bỏ qua dòng header).
        Cú pháp: excel stat <cột>
        Hiển thị: sum, avg, min, max, count, số lượng null
        """
        if not args:
            print("⚠️ excel stat <cột>")
            return
        col = int(args[0])
        values = []
        null_count = 0
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is None:
                null_count += 1
            else:
                try:
                    values.append(float(val))
                except (ValueError, TypeError):
                    pass  # bỏ qua giá trị không phải số
        if not values:
            print("⚠️ Không có dữ liệu số trong cột")
            return
        total = sum(values)
        avg_val = total / len(values)
        min_val = min(values)
        max_val = max(values)
        count = len(values)
        print(f"📊 Thống kê cột {col}:")
        print(f"   Tổng: {total}")
        print(f"   TB  : {avg_val:.2f}")
        print(f"   Min : {min_val}")
        print(f"   Max : {max_val}")
        print(f"   Số lượng số: {count}")
        print(f"   Ô trống: {null_count}")
        # Lưu vào context
        self.assistant.context[f'stat_col_{col}'] = {'sum': total, 'avg': avg_val, 'min': min_val, 'max': max_val}
    
    
    @with_worksheet
    def cmd_transpose(self, ws, args):
        """
        Chuyển vị dữ liệu (hàng thành cột, cột thành hàng).
        Cú pháp: excel transpose
        """
        max_row = ws.max_row
        max_col = ws.max_column
        if max_row == 0 or max_col == 0:
            print("⚠️ Không có dữ liệu để chuyển vị")
            return
        # Đọc toàn bộ dữ liệu
        data = []
        for r in range(1, max_row + 1):
            row_data = []
            for c in range(1, max_col + 1):
                row_data.append(ws.cell(row=r, column=c).value)
            data.append(row_data)
        # Xóa nội dung cũ
        ws.delete_rows(1, ws.max_row)
        # Ghi dữ liệu đã chuyển vị
        for new_c in range(1, len(data) + 1):
            for new_r in range(1, len(data[0]) + 1):
                ws.cell(row=new_r, column=new_c).value = data[new_c-1][new_r-1]
        print(f"✅ Đã chuyển vị ma trận {max_row}x{max_col} → {max_col}x{max_row}")
    
    
    @with_worksheet
    def cmd_merge_sheets(self, ws, args):
        """
        Hợp nhất dữ liệu từ nhiều sheet trong cùng workbook vào sheet hiện tại.
        Cú pháp: excel merge_sheets sheet2 sheet3 ...
        """
        if not args:
            print("⚠️ excel merge_sheets <tên_sheet1> <tên_sheet2> ...")
            return
        wb = load_workbook(self.file)  # lấy lại workbook gốc
        target_ws = wb.active  # sheet đang active
        header = None
        # Dòng bắt đầu ghi (giả sử sheet hiện tại có thể có header)
        current_row = target_ws.max_row + 1 if target_ws.max_row > 0 else 1
        
        for sheet_name in args:
            if sheet_name not in wb.sheetnames:
                print(f"⚠️ Sheet '{sheet_name}' không tồn tại, bỏ qua")
                continue
            src_ws = wb[sheet_name]
            # Nếu là sheet đầu tiên và chưa có header thì lấy header từ sheet đó
            if header is None and current_row == 1 and target_ws.max_row == 0:
                header = [src_ws.cell(row=1, column=c).value for c in range(1, src_ws.max_column + 1)]
                for col_idx, val in enumerate(header, start=1):
                    target_ws.cell(row=1, column=col_idx).value = val
                current_row = 2
                # Bỏ qua dòng header của source
                start_src_row = 2
            else:
                start_src_row = 1
            # Copy dữ liệu từ source sang target
            for r in range(start_src_row, src_ws.max_row + 1):
                for c in range(1, src_ws.max_column + 1):
                    val = src_ws.cell(row=r, column=c).value
                    target_ws.cell(row=current_row, column=c).value = val
                current_row += 1
        wb.save(self.file)
        print(f"✅ Đã hợp nhất {len(args)} sheet vào sheet hiện tại")

    @with_worksheet
    def cmd_formula(self, ws, args):
        """
        Gán công thức cho một ô.
        Cú pháp: excel formula <hàng> <cột> <công_thức>
        Ví dụ: excel formula 2 3 "=A1+B1"
               excel formula 1 1 "SUM(A2:A10)"
        """
        if len(args) < 3:
            print("⚠️ excel formula <hàng> <cột> <công_thức>")
            return
        try:
            row = int(args[0])
            col = int(args[1])
        except ValueError:
            print("⚠️ Hàng và cột phải là số")
            return
    
        # Ghép lại phần còn lại thành công thức (có thể chứa dấu cách)
        formula = ' '.join(args[2:]).strip()
        if not formula.startswith('='):
            formula = '=' + formula
    
        ws.cell(row=row, column=col).value = formula
        print(f"✅ Đã gán công thức '{formula}' vào ô ({row},{col})")

def register(assistant):
    assistant.handlers.append(ExcelProHandler(assistant))

plugin_info = {
    'enabled': True,
    'register': register,
    'command_handle': [
        'excel -f data.xlsx create',
        'excel -f data.xlsx add 10 20 30',
        'excel -f data.xlsx read',
        'excel -f data.xlsx update 2 1 hello',
        'excel -f data.xlsx delete 2',
        'excel -f data.xlsx delrows 1 10',
        'excel -f data.xlsx delcolrange 1 2 5',
        'excel -f data.xlsx find 20',
        'excel -f data.xlsx avg',
        'excel -f data.xlsx avg_range 1 2 4',
        'excel -f data.xlsx chart',
        'excel -f data.xlsx auto',
        'excel -f data.xlsx header 1 "Mã số"',
        'excel -f data.xlsx header "Tên" "Tuổi" "Điểm"',
        'excel -f data.xlsx comment 2 3 "Đây là ghi chú"',
        'excel -f data.xlsx manual',
        'excel setfile myfile.xlsx',
        'excel copy backup.xlsx',
        'excel copy source.xlsx dest.xlsx',
        'excel -f data.xlsx autofit',           # tự động tất cả
        'excel -f data.xlsx autofit 1 3 5',     # chỉ cột 1,3,5
        'excel -f data.xlsx colorminmax 2',   # tô màu min/max ở cột số 2
            # Tìm và thay thế
    'excel -f data.xlsx find_replace "old" "new"'
    
    # Xóa dòng trống
    'excel -f data.xlsx remove_empty_rows'
    
    # Sắp xếp cột 2 (tăng dần)
    'excel -f data.xlsx sort 2 asc'
    
    # Thống kê cột 3
    'excel -f data.xlsx stat 3'
    
    # Chuyển vị dữ liệu
    'excel -f data.xlsx transpose'
    
    # Hợp nhất dữ liệu từ các sheet "Sales" và "Inventory" vào sheet hiện tại
    'excel -f data.xlsx merge_sheets Sales Inventory'
        'excel -f data.xlsx formula 2 3 "=A1+B1"', # công thức cho ô
    'excel -f data.xlsx formula 4 1 "SUM(A2:A3)"', # công thức cho ô


    ],
}