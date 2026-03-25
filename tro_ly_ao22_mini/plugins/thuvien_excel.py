# plugins/excel_plugin.py
import os
from typing import Any
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system("pip install openpyxl")
    import openpyxl
    from openpyxl.utils import get_column_letter

class ExcelHandler:
    def __init__(self):
        self.can_handle_commands = [
            "list sheets", "read sheet", "read cell",
            "write cell", "add row", "create file",
            "delete sheet", "delete file"
        ]
        self.loaded_workbooks = {}  # Lưu workbook đã mở theo đường dẫn

    def can_handle(self, command: str) -> bool:
        return command.lower() in self.can_handle_commands

    def handle(self, command: str) -> None:
        cmd = command.lower()
        if cmd == "list sheets":
            self.list_sheets()
        elif cmd == "read sheet":
            self.read_sheet()
        elif cmd == "read cell":
            self.read_cell()
        elif cmd == "write cell":
            self.write_cell()
        elif cmd == "add row":
            self.add_row()
        elif cmd == "create file":
            self.create_file()
        elif cmd == "delete sheet":
            self.delete_sheet()
        elif cmd == "delete file":
            self.delete_file()

    # ======= Các chức năng =======
    def _get_workbook(self, path: str):
        if path in self.loaded_workbooks:
            return self.loaded_workbooks[path]
        if not os.path.exists(path):
            print("⚠️ File không tồn tại!")
            return None
        try:
            wb = openpyxl.load_workbook(path)
            self.loaded_workbooks[path] = wb
            return wb
        except Exception as e:
            print(f"⚠️ Lỗi khi mở file Excel: {e}")
            return None

    def list_sheets(self):
        path = input("Nhập đường dẫn file Excel: ").strip()
        wb = self._get_workbook(path)
        if wb:
            print("✅ Các sheet trong file:")
            for sheet in wb.sheetnames:
                print(f" - {sheet}")

    def read_sheet(self):
        path = input("Nhập đường dẫn file Excel: ").strip()
        sheet_name = input("Nhập tên sheet: ").strip()
        wb = self._get_workbook(path)
        if wb:
            if sheet_name not in wb.sheetnames:
                print("⚠️ Sheet không tồn tại!")
                return
            sheet = wb[sheet_name]
            print(f"✅ Nội dung sheet {sheet_name}:")
            for row in sheet.iter_rows(values_only=True):
                print(row)

    def read_cell(self):
        path = input("Nhập đường dẫn file Excel: ").strip()
        sheet_name = input("Nhập tên sheet: ").strip()
        cell = input("Nhập ô (ví dụ: A1): ").strip().upper()
        wb = self._get_workbook(path)
        if wb:
            if sheet_name not in wb.sheetnames:
                print("⚠️ Sheet không tồn tại!")
                return
            sheet = wb[sheet_name]
            try:
                value = sheet[cell].value
                print(f"✅ Giá trị ô {cell}: {value}")
            except Exception as e:
                print(f"⚠️ Lỗi khi đọc ô: {e}")

    def write_cell(self):
        path = input("Nhập đường dẫn file Excel: ").strip()
        sheet_name = input("Nhập tên sheet: ").strip()
        cell = input("Nhập ô muốn ghi (ví dụ: B2): ").strip().upper()
        value = input("Nhập giá trị: ").strip()
        wb = self._get_workbook(path)
        if wb:
            if sheet_name not in wb.sheetnames:
                print("⚠️ Sheet không tồn tại!")
                return
            sheet = wb[sheet_name]
            sheet[cell] = value
            wb.save(path)
            print(f"✅ Đã ghi giá trị '{value}' vào ô {cell}")

    def add_row(self):
        path = input("Nhập đường dẫn file Excel: ").strip()
        sheet_name = input("Nhập tên sheet: ").strip()
        wb = self._get_workbook(path)
        if wb:
            if sheet_name not in wb.sheetnames:
                print("⚠️ Sheet không tồn tại!")
                return
            sheet = wb[sheet_name]
            print("Nhập dữ liệu cho dòng mới, cách nhau bằng dấu ,")
            row_data = input("Dòng: ").strip().split(",")
            sheet.append([cell.strip() for cell in row_data])
            wb.save(path)
            print("✅ Đã thêm dòng vào sheet")

    def create_file(self):
        path = input("Nhập tên file Excel muốn tạo: ").strip()
        sheet_name = input("Nhập tên sheet mặc định: ").strip() or "Sheet1"
        wb = openpyxl.Workbook()
        wb.active.title = sheet_name
        wb.save(path)
        print(f"✅ Đã tạo file Excel: {path}")

    def delete_sheet(self):
        path = input("Nhập đường dẫn file Excel: ").strip()
        sheet_name = input("Nhập tên sheet muốn xóa: ").strip()
        wb = self._get_workbook(path)
        if wb:
            if sheet_name not in wb.sheetnames:
                print("⚠️ Sheet không tồn tại!")
                return
            del wb[sheet_name]
            wb.save(path)
            print(f"✅ Đã xóa sheet {sheet_name}")

    def delete_file(self):
        path = input("Nhập đường dẫn file Excel muốn xóa: ").strip()
        if os.path.exists(path):
            os.remove(path)
            print(f"✅ Đã xóa file {path}")
        else:
            print("⚠️ File không tồn tại!")

# ===== Thông tin plugin =====
plugin_info = {
    "enabled": True,
    "register": lambda assistant: assistant.handlers.append(ExcelHandler()),
    "command_handle": [
        "list sheets", "read sheet", "read cell",
        "write cell", "add row", "create file",
        "delete sheet", "delete file"
    ]
}