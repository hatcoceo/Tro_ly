import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


class ExcelBangVatTuCreator:

    def __init__(self):
        self.filename = 'bang_vat_tu.xlsx'

    def tao_file_excel(self):
        wb = Workbook()
        ws_vattu = wb.active
        ws_vattu.title = 'vat_tu'
        ws_dongia = wb.create_sheet('don_gia')
        ws_dongia.append(['Tên hàng', 'Đơn giá'])
        ws_dongia.append(['Gạch A', 120000])
        ws_dongia.append(['Gạch B', 100000])
        ws_dongia.append(['Gạch C', 80000])
        ws_dongia.append(['Gạch D', 60000])
        headers = ['Tên hàng', 'Dài', 'Rộng', 'Diện tích', 'Số lượng',
            'Đơn giá', 'Thành tiền', 'Ghi chú']
        ws_vattu.append(headers)
        for col in ws_vattu.iter_cols(min_row=1, max_row=1):
            for cell in col:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
        for i in range(2, 22):
            dai_cell = f'B{i}'
            rong_cell = f'C{i}'
            dientich_cell = f'D{i}'
            soluong_cell = f'E{i}'
            dongia_cell = f'F{i}'
            thanhtien_cell = f'G{i}'
            tenhang_cell = f'A{i}'
            ws_vattu[dientich_cell] = f'={dai_cell}*{rong_cell}'
            ws_vattu[dongia_cell
                ] = f'=IFERROR(VLOOKUP({tenhang_cell},don_gia!A:B,2,FALSE),0)'
            ws_vattu[thanhtien_cell
                ] = f'={dientich_cell}*{dongia_cell}*{soluong_cell}'
        wb.save(self.filename)
        print(f'✅ Đã tạo file Excel: {self.filename}')


def register(assistant):
    creator = ExcelBangVatTuCreator()


    class TaoExcelCommand:

        def can_handle(self, command: str) ->bool:
            return command.strip().lower() == 'tạo bảng vật tư'

        def handle(self, command: str) ->bool:
            creator.tao_file_excel()
            return True
    assistant.handlers.insert(2, TaoExcelCommand())


plugin_info = {'enabled': False, 'register': register, 'classes': [],
    'methods': []}
