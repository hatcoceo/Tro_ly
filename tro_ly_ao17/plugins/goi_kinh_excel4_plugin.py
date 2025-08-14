import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

class ExcelBangVatTuCreator:
    def __init__(self):
        self.filename = "bang_vat_tu2.xlsx"

    def tao_file_excel(self):
        wb = Workbook()
        ws_vattu = wb.active
        ws_vattu.title = "vat_tu"

        # Sheet đơn giá
        ws_dongia = wb.create_sheet("don_gia")
        ws_dongia.append(["Tên hàng", "Đơn giá"])
        ws_dongia.append(["Gạch A", 120000])
        ws_dongia.append(["Gạch B", 100000])
        ws_dongia.append(["Gạch C", 80000])
        ws_dongia.append(["Gạch D", 60000])

        # Sheet vật tư với công thức liên kết
        headers = ["Tên hàng", "Dài", "Rộng", "Diện tích", "Số lượng", "Đơn giá", "Thành tiền", "Ghi chú"]
        ws_vattu.append(headers)

        for col in ws_vattu.iter_cols(min_row=1, max_row=1):
            for cell in col:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

        # Tạo sẵn 20 dòng mẫu với công thức
        for i in range(2, 22):
            dai_cell = f"B{i}"
            rong_cell = f"C{i}"
            dientich_cell = f"D{i}"
            soluong_cell = f"E{i}"
            dongia_cell = f"F{i}"
            thanhtien_cell = f"G{i}"
            tenhang_cell = f"A{i}"

            ws_vattu[dientich_cell] = f"={dai_cell}*{rong_cell}"
            ws_vattu[dongia_cell] = f"=IFERROR(VLOOKUP({tenhang_cell},don_gia!A:B,2,FALSE),0)"
            ws_vattu[thanhtien_cell] = f"={dientich_cell}*{dongia_cell}*{soluong_cell}"

        wb.save(self.filename)
        print(f"✅ Đã tạo file Excel: {self.filename}")

    def reset_so_luong(self):
        wb = openpyxl.load_workbook(self.filename)
        ws_vattu = wb["vat_tu"]

        max_row = ws_vattu.max_row
        for row in range(2, max_row + 1):
            ws_vattu[f"E{row}"] = 0

        wb.save(self.filename)
        print("🔄 Đã đặt lại tất cả số lượng về 0.")

    def tao_sheet_con_hang(self):
        wb = openpyxl.load_workbook(self.filename)
        ws_vattu = wb["vat_tu"]

        # Xóa sheet cũ nếu tồn tại
        if "con_hang" in wb.sheetnames:
            del wb["con_hang"]

        ws_conhang = wb.create_sheet("con_hang")
        ws_conhang.append(["Tên hàng", "Số lượng"])

        for row in ws_vattu.iter_rows(min_row=2, values_only=True):
            ten_hang, _, _, _, so_luong = row[:5]
            if ten_hang and isinstance(so_luong, (int, float)) and so_luong > 0:
                ws_conhang.append([ten_hang, so_luong])

        wb.save(self.filename)
        print("📄 Đã cập nhật sheet 'con_hang' với các hàng còn số lượng > 0.")

# ==================== PLUGIN INFO ====================

def register(assistant):
    creator = ExcelBangVatTuCreator()

    class TaoExcelCommand:
        def can_handle(self, command: str) -> bool:
            return command.strip().lower() == "tạo bảng vật tư"

        def handle(self, command: str) -> bool:
            creator.tao_file_excel()
            creator.tao_sheet_con_hang()  # Tự động tạo sheet con_hang
            return True

    class ResetSoLuongCommand:
        def can_handle(self, command: str) -> bool:
            return command.strip().lower() == "reset số lượng"

        def handle(self, command: str) -> bool:
            creator.reset_so_luong()
            creator.tao_sheet_con_hang()  # Tự động cập nhật sheet con_hang
            return True

    assistant.handlers.insert(2, TaoExcelCommand())
    assistant.handlers.insert(3, ResetSoLuongCommand())

plugin_info = {
    "enabled": True,
    "register": register,
    "classes": [],
    "methods": []
}