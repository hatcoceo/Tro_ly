from tabulate import tabulate
import os
from openpyxl import Workbook
import openpyxl
from typing import List, Dict, Any

plugin_info = {
    "name": "Excel Utility Library",
    "version": "1.0",
    "enabled": True,
    "register": lambda assistant: register_excel_utils(assistant)
}

def register_excel_utils(assistant):
    assistant.context["excel_utils"] = ExcelUtils()
    #print("📚 Excel utility library loaded into context as 'excel_utils'")


class ExcelUtils:
    def create_sample_workbook(self, file_path: str = "data.xlsx"):
        """Tạo file Excel mẫu với dữ liệu tên và tuổi"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["Tên", "Tuổi", "Địa Chỉ" ])
        ws.append(["An", 25, "Lê Chân"])
        ws.append(["Bình", 30, "Lê Duẩn"])
        wb.save(file_path)
        print(f"📁 Đã tạo file Excel mẫu tên: {file_path}")    
    
    def load_workbook(self, file_path: str):
        """Load an Excel workbook"""
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File không tồn tại: {file_path}")
        return openpyxl.load_workbook(file_path)
    
    def save_workbook(self, workbook, file_path: str):
        """Save an Excel workbook"""
        workbook.save(file_path)

    def get_sheet_names(self, workbook) -> List[str]:
        """Get all sheet names in a workbook"""
        return workbook.sheetnames

    def get_sheet(self, workbook, sheet_name: str):
        """Get a specific sheet by name"""
        return workbook[sheet_name]

    def read_rows(self, sheet, min_row=1, max_row=None) -> List[List[Any]]:
        """Read rows from sheet"""
        data = []
        for row in sheet.iter_rows(min_row=min_row, max_row=max_row, values_only=True):
            data.append(list(row))
        return data

    def write_row(self, sheet, values: List[Any]):
        """Append a row to the sheet"""
        sheet.append(values)

    def update_cell(self, sheet, row: int, col: int, value: Any):
        """Update a specific cell"""
        sheet.cell(row=row, column=col, value=value)

    def delete_row(self, sheet, row: int):
        """Delete a row from the sheet"""
        sheet.delete_rows(row)

    def find_rows_by_value(self, sheet, column_index: int, value: Any) -> List[int]:
        """Find row indices that contain a value in a given column (1-based)"""
        matching_rows = []
        for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if len(row) >= column_index and row[column_index - 1] == value:
                matching_rows.append(i)
        return matching_rows
    
    def get_cell_value(self, sheet, row: int, col: int) -> Any:
        """Get value from specific cell"""
        return sheet.cell(row=row, column=col).value
    
    def find_rows_by_conditions(self, sheet, conditions: Dict[int, Any]) -> List[int]:
        """
        Tìm các dòng thỏa nhiều điều kiện trên các cột (dùng index cột 1-based)
        conditions = {1: "An", 2: 25}
        """
        matching = []
        for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if all(len(row) >= col_idx and row[col_idx - 1] == val for col_idx, val in conditions.items()):
                matching.append(i)
        return matching
    
    def delete_rows(self, sheet, rows: List[int]):
        """
        Xoá nhiều dòng (chú ý: phải xóa từ cuối lên để không bị lệch chỉ số)
        """
        for row in sorted(rows, reverse=True):
            sheet.delete_rows(row)
    
    def overwrite_sheet(self, sheet, data: List[List[Any]]):
        """
        Ghi đè toàn bộ nội dung sheet với dữ liệu mới
        """
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            sheet.delete_rows(row[0].row)
        for row in data:
            sheet.append(row)
    
    def create_workbook_with_sheets(self, sheets_data: Dict[str, List[List[Any]]], file_path: str):
        """
        Tạo file Excel với nhiều sheet, mỗi sheet có data riêng
        """
        wb = Workbook()
        first = True
        for name, data in sheets_data.items():
            ws = wb.active if first else wb.create_sheet(title=name)
            ws.title = name
            for row in data:
                ws.append(row)
            first = False
        wb.save(file_path)
    def print_rows(self, data: List[List[Any]], headers: bool = True):
        """
        In dữ liệu hàng loạt ra màn hình dưới dạng bảng
        """
        if not data:
            print("⚠️ Không có dữ liệu để hiển thị.")
            return

        if headers:
            table = tabulate(data[1:], headers=data[0], tablefmt="grid")
        else:
            table = tabulate(data, tablefmt="fancy_grid")
        print(table)

    def print_sheet(self, sheet, min_row=1, max_row=None):
        """
        In nội dung một sheet ra màn hình
        """
        data = self.read_rows(sheet, min_row=min_row, max_row=max_row)
        self.print_rows(data)
    
    def set_formula(self, sheet, row: int, col: int, formula: str):
        """
        Gán công thức vào ô (ví dụ: "=SUM(A1:A10)")
        """
        sheet.cell(row=row, column=col).value = f"={formula}"
    
    
    def apply_sum_formula_to_column(self, sheet, col: int, start_row: int, end_row: int, target_row: int):
        """
        Tạo công thức SUM cho một cột (1-based index) và gán vào dòng đích.
        Ví dụ: =SUM(B2:B10) tại ô B11
        """
        col_letter = openpyxl.utils.get_column_letter(col)
        formula = f"SUM({col_letter}{start_row}:{col_letter}{end_row})"
        self.set_formula(sheet, target_row, col, formula)
    
    def apply_average_formula_to_column(self, sheet, col: int, start_row: int, end_row: int, target_row: int):
        """
        Gán công thức AVERAGE vào một ô
        """
        col_letter = openpyxl.utils.get_column_letter(col)
        formula = f"AVERAGE({col_letter}{start_row}:{col_letter}{end_row})"
        self.set_formula(sheet, target_row, col, formula)
    
    def apply_custom_formula(self, sheet, formula_template: str, row: int, col: int):
        """
        Gán công thức tùy chỉnh (ví dụ: =A1+B1 hoặc =IF(A2>10, "Yes", "No"))
        """
        self.set_formula(sheet, row, col, formula_template)
    
    def add_total_row(self, sheet, start_row: int, end_row: int, target_row: int, exclude_columns: List[int] = []):
        """
        Tự động thêm dòng tổng cho các cột, trừ cột bị loại trừ (1-based).
        """
        for col in range(1, sheet.max_column + 1):
            if col in exclude_columns:
                continue
            try:
                float(sheet.cell(row=start_row, column=col).value or 0)
                self.apply_sum_formula_to_column(sheet, col, start_row, end_row, target_row)
            except ValueError:
                pass  # Bỏ qua nếu không phải số
