# xuất txt sang excel tri_thuc.txt => data/tri_thuc.xlsx
#  xuất txt sang excel tri_thuc.txt => tri_thuc.xlsx
# xuất txt sang excel data/tri_thuc2.txt => data/tri_thuc2.xlsx
import os
import openpyxl
from openpyxl.utils import get_column_letter

class TxtToExcelExporter:
    def export(self, txt_path: str, excel_path: str):
        if not os.path.exists(txt_path):
            print(f"❌ File không tồn tại: {txt_path}")
            return False

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tri Thuc"

        with open(txt_path, "r", encoding="utf-8") as f:
            lines = f.readlines()

        ws.append(["Câu hỏi", "Câu trả lời"])
        for line in lines:
            if "||" in line:
                parts = line.strip().split("||", 1)
                question = parts[0].strip()
                answer = parts[1].strip()
                ws.append([question, answer])
            else:
                print(f"⚠️ Dòng không hợp lệ (bỏ qua): {line.strip()}")

        # Tự động điều chỉnh độ rộng cột
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        wb.save(excel_path)
        print(f"✅ Xuất dữ liệu thành công ra: {excel_path}")
        return True

# Plugin handler
class TxtToExcelCommandHandler:
    def can_handle(self, command: str) -> bool:
        return command.startswith("xuất txt sang excel")

    def handle(self, command: str) -> bool:
        try:
            parts = command[len("xuất txt sang excel"):].strip().split("=>")
            if len(parts) != 2:
                print("⚠️ Cú pháp đúng: xuất txt sang excel <đường_dẫn_txt> => <đường_dẫn_excel>")
                return True

            txt_path = parts[0].strip()
            excel_path = parts[1].strip()
            exporter = TxtToExcelExporter()
            exporter.export(txt_path, excel_path)
        except Exception as e:
            print(f"❌ Lỗi khi xử lý lệnh: {e}")
        return True

# Plugin info
plugin_info = {
    "enabled": True,
    "methods": [
        {
            "class_name": "TxtToExcelExporter",
            "method_name": "export",
            "version": "v1",
            "function": TxtToExcelExporter().export,
            "description": "Xuất dữ liệu từ file txt sang Excel"
        }
    ],
    "classes": [
        {
            "class_name": "TxtToExcelExporter",
            "version": "v1",
            "class_ref": TxtToExcelExporter
        }
    ],
#    "register": lambda core: core.handlers.append(TxtToExcelCommandHandler())
       "register": lambda core: core.handlers.insert(4, TxtToExcelCommandHandler())
}